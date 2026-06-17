#!/usr/bin/env python3
"""
FAMIC 公開 CSV と CropLife Japan RAC Excel を統合して
ブラウザ用の pesticides.json を生成する。

Data sources:
  - FAMIC 基本部:  R0804080.csv  (Shift_JIS, 11列, 6339行 / 3801商品)
                   ※混合剤は成分ごとに行が分かれている
  - FAMIC 適用部1: R0804081.csv  (Shift_JIS, 25列, 登録番号 52〜22553)
  - FAMIC 適用部2: R0804082.csv  (Shift_JIS, 25列, 登録番号 22554〜)
  - CropLife RAC: mechanism_rac.xlsx  (UTF-8, 製品名順, 約3968件)

Join key: 登録番号 (integer)

Output:
  src/data/pesticides.json       - メインDB（基本情報＋成分＋RAC）、シングルHTML埋め込み用
  src/data/applications.json     - 適用部詳細（対象作物・病害虫等）、オンデマンド読込
"""

import csv
import json
import re
import sys
import unicodedata
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("ERROR: openpyxl が必要です。pip3 install openpyxl を実行してください。")

try:
    import xlrd  # 旧 .xls 用
except ImportError:
    xlrd = None

PROJECT_ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = PROJECT_ROOT / "rawdata"
OUT_DIR = PROJECT_ROOT / "src" / "data"
OUT_MAIN = OUT_DIR / "pesticides.json"
OUT_APPS = OUT_DIR / "applications.json"

def _pick_famic_csv(suffix_digit: str) -> Path:
    """FAMIC CSV は月ごとに R{YYMMDD}{0|1|2}.csv と名前が変わるため、
    末尾1桁 (0=基本部, 1=適用部一, 2=適用部二) で最新を自動検出する。"""
    candidates = sorted((RAW_DIR / "famic").glob(f"R*{suffix_digit}.csv"))
    if not candidates:
        sys.exit(
            f"ERROR: rawdata/famic/R*{suffix_digit}.csv が見つかりません。\n"
            f"       scripts/update_data.py で最新データを取得してください。"
        )
    return candidates[-1]  # アルファベット順 = 日付順の最新


def _famic_date_from_path(path: Path) -> str:
    """R0806100.csv → '2026-06-10'  (令和YY年MM月DD日 → 西暦 ISO 形式)"""
    m = re.match(r"R(\d{2})(\d{2})(\d{2})\d\.csv", path.name)
    if not m:
        return "不明"
    year = 2018 + int(m.group(1))
    return f"{year}-{m.group(2)}-{m.group(3)}"


def _sikkou_date_from_path(path: Path) -> str:
    """sikkounouyaku_20260531.xls → '2026-05-31'"""
    m = re.match(r"sikkounouyaku_(\d{4})(\d{2})(\d{2})\.xls", path.name)
    if not m:
        return "不明"
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"


def _pick_sikkou_xls() -> Path:
    """失効農薬一覧 .xls はファイル名の日付 (sikkounouyaku_YYYYMMDD.xls) が
    更新ごとに変わるため、最新を自動検出する。失効剤データは任意項目なので、
    見つからない場合は存在しないパスを返し、load_cancelled_pesticides() の
    存在チェックでスキップさせる（FAMIC本体と違い sys.exit はしない）。"""
    candidates = sorted((RAW_DIR / "sikkou").glob("sikkounouyaku_*.xls"))
    if not candidates:
        return RAW_DIR / "sikkou" / "sikkounouyaku_not_found.xls"
    return candidates[-1]  # ファイル名末尾が日付なのでアルファベット順 = 最新


FAMIC_BASIC = _pick_famic_csv("0")
FAMIC_APP1 = _pick_famic_csv("1")
FAMIC_APP2 = _pick_famic_csv("2")
RAC_XLSX = RAW_DIR / "mechanism_rac.xlsx"
RAC_MANUAL = RAW_DIR / "rac_manual.json"
SIKKOU_XLS = _pick_sikkou_xls()

TARGET_CATEGORIES = {"殺虫剤", "殺菌剤", "除草剤"}

HOUSEHOLD_KEYWORDS = ("家庭園芸", "ガーデニング", "家庭用")
# FAMIC 剤型名の表記ゆれ: エアゾル / エアロゾル / スプレー
HOUSEHOLD_TYPE_KEYWORDS = ("エアゾル", "エアロゾル", "スプレー")


def norm(text):
    """半角カタカナ → 全角、トリム。"""
    if text is None:
        return ""
    return unicodedata.normalize("NFKC", str(text).strip())


def pick_categories(yoto):
    """
    用途文字列から該当する全カテゴリを返す（複数カテゴリ対応）。
    例: "殺虫殺菌剤" → ["殺虫剤", "殺菌剤"]
        "殺虫除草剤" → ["殺虫剤", "除草剤"]
        "殺菌植調剤" → ["殺菌剤"]  (植調は対象外)
    """
    y = norm(yoto)
    cats = []
    if "殺虫" in y:
        cats.append("殺虫剤")
    if "殺菌" in y:
        cats.append("殺菌剤")
    if "除草" in y:
        cats.append("除草剤")
    return cats


# 用途名・種類名が FAMIC 表記と実際の作用機序で乖離する剤を上書き分類する。
# キーは type_name (基本部) または original_category 用途欄の文字列。
# 値は ["殺虫剤", "殺菌剤"] のような正規分類。pick_categories で得た結果を上書きする。
TYPE_NAME_CATEGORY_OVERRIDE = {
    # 硫黄系: 殺ダニ・カイガラムシ防除に使われる殺虫殺菌両作用剤
    "石灰硫黄合剤": ["殺虫剤", "殺菌剤"],
    "硫黄合剤": ["殺虫剤", "殺菌剤"],
    "水和硫黄剤": ["殺虫剤", "殺菌剤"],
    # 松脂合剤: カイガラムシ防除を主目的とする殺虫殺菌剤
    "松脂合剤": ["殺虫剤", "殺菌剤"],
    # マシン油剤: 殺ダニ・カイガラムシ防除の殺虫剤として登録 (殺菌作用は副次)
    # ボルドー液: 殺菌剤 (修正なし)
}


def apply_category_override(type_name, categories):
    """type_name が上書きマップに含まれていれば上書き分類を返す。"""
    if not type_name:
        return categories
    if type_name in TYPE_NAME_CATEGORY_OVERRIDE:
        return list(TYPE_NAME_CATEGORY_OVERRIDE[type_name])
    # 末尾一致 (例: "サンケイ石灰硫黄合剤" は商品名なので使わない、type_name のみ対象)
    for key, val in TYPE_NAME_CATEGORY_OVERRIDE.items():
        if type_name == key:
            return list(val)
    return categories


# 失効剤の旧農薬成分名からカテゴリを推定する補助テーブル（殺虫/殺菌/除草 が種類名に含まれない場合用）
# 2026-06-17: 上位頻度の失効剤ルート約150件をWebSearchで個別調査し、確信度高/中で
# 同定できたものを追加（出典: 環境省・FAMIC農薬抄録・Wikipedia等）。線虫防除剤・
# 軟体動物防除剤・殺ダニ専用剤は、現行データ(ピリダベン等)で「殺虫剤」に分類されて
# いる実例を確認した上で「殺虫剤」に含めている。確信を持てなかった項目（PMP, CYP,
# 有機ひ素, BEBP, CVMP, TUZ, CMP, DCPA, CPA, BCPE, チアジアジン, DAPA, ESBP, ETM,
# CPAS, APC, ジクロン, フェナジンオキシド, DPC, ESP, 蛋白加水分解物, EPBP, 粘着,
# CNA, アルカリ, 果実防腐, テミビンホス, DAEP, ベスロジン, ジオキサン系有機りん,
# FABA, マイトメート, 硫酸亜鉛, クロラムフェニコール, サリチオン, DCPA, MBCP, CVP,
# カーバノレート, ホルモチオン, スルフェン酸系, 有機ニッケル, アンスラキノン 等）は
# 追加せず、従来どおり除外を維持する。
LEGACY_CATEGORY_HINTS = {
    "殺虫剤": [
        "DDT", "BHC", "ひ酸鉛", "ヒ酸鉛", "ひ酸石灰", "ヒ酸石灰",
        "パラチオン", "マラチオン", "ディールドリン", "アルドリン", "エンドリン",
        "クロルデン", "リンデン", "ロテノン", "デリス", "ピレトリン", "除虫菊",
        "ニコチン", "エチオン", "ダイアジノン", "カルバリル", "ディプテレックス",
        "グラヤノトキシン", "ハイトロキシド", "シアン化水素", "クロルピクリン",
        "EPN", "MPP", "DEP", "DMTP", "松脂合剤",
        "NAC", "MTMC", "BPMC", "MPMC", "DDVP", "PHC", "XMC",
        "ヘプタクロル", "プロパホス", "ピリダフェンチオン", "シクロプロトリン",
        "ディルドリン", "TEPP", "クロルベンジレート", "CYAP", "シラフルオフェン",
        "エチルチオメトン", "ベンゾエピン", "クロルピリホス", "クロルピリホスメチル",
        "アラマイト", "酸化フェンブタスズ", "フェニソブロモレート", "ホサロン",
        "ベンスルタップ", "ピラクロホス", "メチルジメトン", "チオメトン",
        "ピリミホスメチル", "テロドリン", "モノクロトホス", "MNFA", "メカルバム",
        "クロフェンテジン", "ケルセン", "CPCBS", "クロルプロピレート",
        "DBCP", "EDB", "なめくじ駆除", "浮塵子駆除",
    ],
    "殺菌剤": [
        "ボルドー", "石灰ボルドー", "銅", "硫酸銅", "塩化第二銅",
        "水銀", "酢酸フェニル水銀", "PMA",
        "硫黄", "石灰硫黄", "ダコニール", "チウラム", "TMTD",
        "キャプタン", "マンネブ", "ジネブ", "マンゼブ", "ベノミル",
        "ストレプトマイシン", "オキシテトラサイクリン", "テレ剤",
        "ポリオキシン", "PCNB", "イミノクタジン酢酸塩", "ジクロシメット",
        "フェナリモル", "チアベンダゾール", "フェノキサニル", "ホルムアルデヒド",
        "ジクロメジン", "ダイホルタン", "シクロヘキシミド", "ビンクロゾリン",
        "メチラム", "ピリフェノックス", "ポリカーバメート", "ファーバム",
        "IBP", "EDDP", "ブラストサイジンS",
    ],
    "除草剤": [
        "2,4-D", "2.4-D", "MCP", "MCPA", "MCPB",
        "パラコート", "ジクワット", "ダイカット",
        "アトラジン", "シマジン", "プロパジン",
        "ピクロラム", "ダイカンバ", "ディカンバ",
        "クロルスルフロン", "テブチウロン",
        "ベンタゾン", "DNBP", "ビフェノックス", "ビアラホス", "スルファミン酸塩",
        "グリホサートトリメシウム塩", "PCP", "メフェナセット", "CNP",
        "エンドタール二ナトリウム塩",
    ],
}


# 剤型カテゴリは 7 種 + その他 に簡素化。
# 優先順:
#   1. 商品名末尾が "フロアブル" / "ジャンボ" / "ベイト" → それぞれ {フロアブル剤, ジャンボ剤, ベイト剤}
#   2. 生の剤型 (FAMIC 剤型名列、または失効剤の種類名から抽出) が
#      {水和剤, 粒剤, 水溶剤, 粉剤} に完全一致 → そのまま
#   3. その他

_ALLOWED_EXACT_FORMULATIONS = {"水和剤", "粒剤", "水溶剤", "粉剤"}


def classify_formulation(product_name, raw_formulation):
    """剤型カテゴリ判定。"""
    pn = norm(product_name or "")
    if pn.endswith("フロアブル"):
        return "フロアブル剤"
    if pn.endswith("ジャンボ"):
        return "ジャンボ剤"
    if pn.endswith("ベイト"):
        return "ベイト剤"
    rf = norm(raw_formulation or "")
    # 失効剤の種類名は "エンドリン乳剤" のような合成語なので末尾一致も許容
    for allowed in _ALLOWED_EXACT_FORMULATIONS:
        if rf == allowed or rf.endswith(allowed):
            return allowed
    return "その他"


def build_ingredient_category_hints(basic_records):
    """
    現行剤の用途データから 成分名→カテゴリ集合 の対応表を動的生成する。
    失効剤の種類名/商品名には用途(殺虫/殺菌/除草)が明記されないことが多く、
    LEGACY_CATEGORY_HINTS (戦前〜昭和期の旧成分名のみ収録) では現代の成分名
    (フサライド、ベンスルフロンメチル等) を取り逃して失効剤がサイレントに
    除外される問題があった。現行剤側で成分名↔用途が既に判明しているため、
    それを失効剤のカテゴリ推定に再利用する。
    戻り値: ({成分名: {カテゴリ, ...}}, 成分名を長い順に並べたリスト)
    """
    votes = defaultdict(set)
    for b in basic_records.values():
        cats = apply_category_override(b["type_name"], pick_categories(b["category"]))
        if not cats:
            continue
        for ing in b["ingredients_raw"]:
            name = ing["name"]
            if len(name) >= 2:
                votes[name].update(cats)
    names_by_length = sorted(votes.keys(), key=len, reverse=True)
    return votes, names_by_length


def pick_categories_with_legacy(type_name, product_name, ingredient_hints=None, ingredient_names_by_length=None):
    """
    通常の pick_categories で拾えない失効剤を、成分名ヒントで救済する。
    1. 現行剤から動的生成した成分名ヒント (現代の成分名をカバー)
    2. 旧農薬成分名ヒント (現行剤に存在しない歴史的成分用、LEGACY_CATEGORY_HINTS)
    """
    cats = pick_categories(type_name + " " + product_name)
    if cats:
        return cats
    combined = norm(type_name) + " " + norm(product_name)

    if ingredient_hints and ingredient_names_by_length:
        for name in ingredient_names_by_length:
            if name in combined:
                return sorted(ingredient_hints[name])

    found = []
    for cat, hints in LEGACY_CATEGORY_HINTS.items():
        for hint in hints:
            if norm(hint) in combined:
                found.append(cat)
                break
    return found


def is_household(product_name, formulation):
    pn = norm(product_name)
    fn = norm(formulation)
    if any(k in pn for k in HOUSEHOLD_KEYWORDS):
        return True
    if any(k in fn for k in HOUSEHOLD_TYPE_KEYWORDS):
        return True
    return False


def load_famic_basic():
    """
    基本部 CSV → {登録番号: {product_name, type, company, mix_count, category, date, ingredients_raw: [{name, density}]}}

    1商品（登録番号）につき複数行ある混合剤は、成分を配列にマージする。
    """
    records = {}
    with open(FAMIC_BASIC, encoding="cp932") as f:
        reader = csv.DictReader(f)
        for row in reader:
            reg_raw = row["登録番号"].strip()
            if not reg_raw:
                continue
            try:
                reg_no = int(reg_raw)
            except ValueError:
                continue

            if reg_no not in records:
                records[reg_no] = {
                    "reg_no": reg_no,
                    "type_name": norm(row["農薬の種類"]),
                    "product_name": norm(row["農薬の名称"]),
                    "company": norm(row["登録を有する者の名称"]),
                    "category": norm(row["用途"]),
                    "formulation": norm(row["剤型名"]),
                    "registration_date": norm(row["登録年月日"]),
                    "mix_count": int(row["混合数"] or 1),
                    "ingredients_raw": [],
                }
            records[reg_no]["ingredients_raw"].append(
                {
                    "name": norm(row["有効成分"]),
                    "density": norm(row["濃度"]),
                    "total_usage": norm(row["総使用回数における有効成分"]),
                }
            )
    return records


def load_rac_excel():
    """RAC Excel → ({登録番号: {initial, ingredients}}, {成分名: RACコード})

    第2返値は成分名→RACコードのマスターテーブル。CropLife表に載っている
    全商品から成分単位で抽出し、多数決で確定する。FAMICの新規登録商品が
    RAC表に未収録でも、その商品の成分が別商品で分類されていればフォール
    バックでマッチできる。
    """
    wb = load_workbook(RAC_XLSX, read_only=True, data_only=True)
    ws = wb["RACコード検索表(製品名順2025.06) "]
    records = {}
    ing_votes = defaultdict(lambda: defaultdict(int))  # {成分名: {rac: count}}
    for r in ws.iter_rows(min_row=4, values_only=True):
        reg_no_cell = r[3]
        reg_no = None
        if reg_no_cell:
            try:
                reg_no = int(reg_no_cell)
            except (ValueError, TypeError):
                reg_no = None

        ingredients = []
        for ing_col, rac_col in [(6, 7), (8, 9), (10, 11), (12, 13), (14, 15)]:
            name = r[ing_col]
            rac = r[rac_col]
            if name:
                rac_str = None
                if rac not in (None, "", 0, "0"):
                    s = norm(str(rac))
                    # RAC Excel は未分類を "-" や "「-」" で表す場合あり
                    if s not in ("-", "「-」", "ー", "―", "N/A"):
                        rac_str = s
                n = norm(name)
                ingredients.append({"name": n, "rac_code": rac_str})
                if rac_str:
                    ing_votes[n][rac_str] += 1

        if reg_no is not None:
            records[reg_no] = {
                "initial": norm(r[1]) if r[1] else None,
                "ingredients": ingredients,
            }

    # 多数決で成分名→RACコードを確定
    ing_master = {
        n: max(votes.items(), key=lambda x: x[1])[0]
        for n, votes in ing_votes.items()
    }
    return records, ing_master


def load_famic_applications():
    """適用部 CSV → {登録番号: [{crop, pest, dosage, timing, count, method}, ...]}"""
    apps = defaultdict(list)
    for path in (FAMIC_APP1, FAMIC_APP2):
        with open(path, encoding="cp932") as f:
            reader = csv.DictReader(f)
            for row in reader:
                reg_no = row["登録番号"].strip()
                if not reg_no:
                    continue
                try:
                    key = int(reg_no)
                except ValueError:
                    continue
                apps[key].append(
                    {
                        "crop": norm(row["作物名"]),
                        "place": norm(row["適用場所"]),
                        "pest": norm(row["適用病害虫雑草名"]),
                        "dosage": norm(row["希釈倍数使用量"]),
                        "spray": norm(row["散布液量"]),
                        "timing": norm(row["使用時期"]),
                        "count": norm(row["本剤の使用回数"]),
                        "method": norm(row["使用方法"]),
                    }
                )
    return apps


def load_manual_rac():
    """rac_manual.json → {成分名: {status, reason}}"""
    if not RAC_MANUAL.exists():
        return {}
    with open(RAC_MANUAL, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("ingredients", {})


def _lookup_rac_by_name(name, ing_master):
    """成分名→RACコードのフォールバック検索。完全一致→部分一致の順。"""
    if not name or not ing_master:
        return None
    if name in ing_master:
        return ing_master[name]
    # 部分一致（塩・異性体表記ゆれ等。"グルホシネートPナトリウム塩" ⊃ "グルホシネート" 等）
    for master_name, code in ing_master.items():
        if master_name in name or name in master_name:
            return code
    return None


def _annotate(name, rac_code, density, manual_map):
    """RAC コードが無い成分には手動ラベルを付与。コードがあれば manual は無視。"""
    entry = {"name": name, "rac_code": rac_code, "density": density}
    if not rac_code:
        man = (manual_map or {}).get(name)
        if man:
            entry["rac_status"] = man.get("status")
            entry["rac_reason"] = man.get("reason", "")
    return entry


def merge_ingredients(basic_rec, rac_rec, ing_master=None, manual_map=None):
    """RAC があれば RAC の成分 + 基本部の濃度を紐づける。RAC 無し/コード無しは
    成分名フォールバック、それでも無ければ手動ラベル (rac_manual.json) を付与。"""
    if rac_rec and rac_rec["ingredients"]:
        out = []
        basic_map = {ing["name"]: ing for ing in basic_rec["ingredients_raw"]}
        for rac_ing in rac_rec["ingredients"]:
            name = rac_ing["name"]
            density = ""
            if name in basic_map:
                density = basic_map[name].get("density", "")
            else:
                for bn, b in basic_map.items():
                    if name in bn or bn in name:
                        density = b.get("density", "")
                        break
            code = rac_ing["rac_code"] or _lookup_rac_by_name(name, ing_master)
            out.append(_annotate(name, code, density, manual_map))
        return out

    # RAC 無し: 基本部の成分 + 成分名フォールバック + 手動ラベル
    return [
        _annotate(
            ing["name"],
            _lookup_rac_by_name(ing["name"], ing_master),
            ing.get("density", ""),
            manual_map,
        )
        for ing in basic_rec["ingredients_raw"]
        if ing["name"]
    ]


def excel_serial_to_date(serial):
    """Excel シリアル日付 → 'YYYY/M/D' 文字列"""
    if not serial or serial == 0:
        return ""
    try:
        from datetime import datetime, timedelta
        # Excel epoch は 1899-12-30 (1900-02-29 バグ補正済みベース)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=int(serial))
        return f"{dt.year}/{dt.month}/{dt.day}"
    except Exception:
        return str(serial)


def load_cancelled_pesticides():
    """
    失効農薬 .xls → List[{reg_no, type_name, product_name, company, reg_date, expire_date, reason}]
    シート「失効農薬一覧」: 登録番号/農薬の種類/農薬の名称/登録を有していた者の略称/登録年月日/失効年月日/失効理由
    シート「失効理由」: 表示番号 → 失効理由テキスト
    """
    if not SIKKOU_XLS.exists() or xlrd is None:
        print("  (失効農薬ファイル無し or xlrd 未インストール → スキップ)")
        return []

    wb = xlrd.open_workbook(str(SIKKOU_XLS))

    # 失効理由マスター
    reason_map = {}
    if "失効理由" in wb.sheet_names():
        sh = wb.sheet_by_name("失効理由")
        for r in range(2, sh.nrows):
            key = sh.cell_value(r, 0)
            val = sh.cell_value(r, 1)
            if key and val:
                try:
                    reason_map[int(key)] = str(val).strip()
                except (ValueError, TypeError):
                    pass

    # 失効農薬一覧
    sh = wb.sheet_by_name("失効農薬一覧")
    records = []
    # ヘッダは row 1 (0-indexed)、row 0 はファイル識別子
    for r in range(2, sh.nrows):
        reg_raw = sh.cell_value(r, 0)
        if not reg_raw:
            continue
        try:
            reg_no = int(reg_raw)
        except (ValueError, TypeError):
            continue

        reason_raw = sh.cell_value(r, 6)
        reason_text = ""
        if reason_raw:
            try:
                reason_text = reason_map.get(int(reason_raw), str(reason_raw).strip())
            except (ValueError, TypeError):
                reason_text = str(reason_raw).strip()

        records.append({
            "reg_no": reg_no,
            "type_name": norm(sh.cell_value(r, 1)),
            "product_name": norm(sh.cell_value(r, 2)),
            "company": norm(sh.cell_value(r, 3)),
            "reg_date": excel_serial_to_date(sh.cell_value(r, 4)),
            "expire_date": excel_serial_to_date(sh.cell_value(r, 5)),
            "reason": reason_text,
        })
    return records


def build():
    print("Loading FAMIC 基本部...")
    basic = load_famic_basic()
    print(f"  基本部: {len(basic)} 商品")

    print("Loading RAC Excel...")
    rac, ing_master = load_rac_excel()
    print(f"  RAC: {len(rac)} 商品 / 成分マスター {len(ing_master)} エントリ")

    print("Loading 手動 RAC ラベル...")
    manual_map = load_manual_rac()
    print(f"  手動ラベル: {len(manual_map)} 成分")

    print("Loading FAMIC 適用部...")
    apps = load_famic_applications()
    total_app_rows = sum(len(v) for v in apps.values())
    print(f"  適用部: {len(apps)} 商品 / 合計 {total_app_rows} 適用エントリ")

    print("Loading 失効農薬一覧...")
    cancelled_raw = load_cancelled_pesticides()
    print(f"  失効: {len(cancelled_raw)} 剤")

    ingredient_hints, ingredient_names_by_length = build_ingredient_category_hints(basic)
    print(f"  失効剤カテゴリ推定用 成分名ヒント: {len(ingredient_hints)} 成分 (現行剤から動的生成)")

    print("\nMerging...")
    products = []
    applications_by_reg = {}
    category_counts = defaultdict(int)
    household_count = 0
    rac_match_count = 0
    cancelled_count = 0

    for reg_no, b in sorted(basic.items()):
        categories = pick_categories(b["category"])
        categories = apply_category_override(b["type_name"], categories)
        if not categories:
            continue  # 殺そ剤・植物成長調整剤など除外

        household = is_household(b["product_name"], b["formulation"])
        if household:
            household_count += 1

        rac_rec = rac.get(reg_no)
        ingredients = merge_ingredients(b, rac_rec, ing_master, manual_map)
        # RAC コードが1つでも付いた商品をマッチとカウント（登録番号 OR 成分名フォールバック）
        if any(ing.get("rac_code") for ing in ingredients):
            rac_match_count += 1

        product = {
            "reg_no": reg_no,
            "product_name": b["product_name"],
            "company": b["company"],
            "type_name": b["type_name"],
            "formulation": classify_formulation(b["product_name"], b["formulation"]),
            "categories": categories,  # 複数カテゴリ配列 (殺虫殺菌剤等に対応)
            "original_category": b["category"],
            "mix_count": b["mix_count"],
            "ingredients": ingredients,
            "household": household,
            "status": "有効",  # FAMIC 公開 CSV は現行登録のみ
            "registration_date": b["registration_date"],
            "app_count": len(apps.get(reg_no, [])),  # 適用エントリ数（詳細読込判定用）
        }
        products.append(product)
        for c in categories:
            category_counts[c] += 1

        if reg_no in apps:
            applications_by_reg[str(reg_no)] = apps[reg_no]

    # 失効剤を追加（成分情報無し、種類名＋商品名＋成分名ヒントからカテゴリ推定）
    active_reg_nos = {p["reg_no"] for p in products}
    cancelled_dropped = []  # カテゴリ推定不能で除外した剤 (サイレント失敗防止のため記録)
    for c in cancelled_raw:
        if c["reg_no"] in active_reg_nos:
            continue  # 現役剤と重複はスキップ

        cats = pick_categories_with_legacy(
            c["type_name"], c["product_name"], ingredient_hints, ingredient_names_by_length
        )
        cats = apply_category_override(c["type_name"], cats)
        if not cats:
            # 殺虫/殺菌/除草 のいずれにも属さないもの（殺そ剤、植調剤等）に加え、
            # ヒントが無く分類不能なものもここに来る。後者を見逃さないよう記録して出力する。
            cancelled_dropped.append(c)
            continue

        household = is_household(c["product_name"], c["type_name"])
        if household:
            household_count += 1

        product = {
            "reg_no": c["reg_no"],
            "product_name": c["product_name"],
            "company": c["company"],
            "type_name": c["type_name"],
            "formulation": classify_formulation(c["product_name"], c["type_name"]),
            "categories": cats,
            "original_category": c["type_name"],
            "mix_count": 1,
            "ingredients": [],  # 失効剤は成分詳細データなし
            "household": household,
            "status": "失効",
            "registration_date": c["reg_date"],
            "expire_date": c["expire_date"],
            "expire_reason": c["reason"],
            "app_count": 0,
        }
        products.append(product)
        cancelled_count += 1
        for cat in cats:
            category_counts[cat] += 1

    today = datetime.now().strftime("%Y-%m-%d")
    famic_date = _famic_date_from_path(FAMIC_BASIC)
    sikkou_date = _sikkou_date_from_path(SIKKOU_XLS) if SIKKOU_XLS.exists() else "不明"

    # メインDB（埋め込み用）
    main_output = {
        "schema_version": "1.0",
        "generated_at": today,
        "sources": {
            "famic": famic_date,
            "famic_cancelled": sikkou_date,
            "rac": "2025-06 (CropLife Japan)",
        },
        "stats": {
            "total_products": len(products),
            "active": len(products) - cancelled_count,
            "cancelled": cancelled_count,
            "by_category": dict(category_counts),
            "household_excluded_by_default": household_count,
            "rac_matched": rac_match_count,
            "rac_match_rate": round(rac_match_count / (len(products) - cancelled_count), 3) if (len(products) - cancelled_count) else 0,
            "manual_labeled_ingredients": len(manual_map),
            "rac_classification_total": sum(
                1 for p in products
                if p.get("status") != "失効"
                and all(
                    ing.get("rac_code") or ing.get("rac_status")
                    for ing in p["ingredients"]
                )
            ),
        },
        "products": products,
    }

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUT_MAIN, "w", encoding="utf-8") as f:
        json.dump(main_output, f, ensure_ascii=False, separators=(",", ":"))
    main_kb = OUT_MAIN.stat().st_size / 1024

    # 適用部（別ファイル、オンデマンド読込用）
    apps_output = {
        "schema_version": "1.0",
        "generated_at": today,
        "applications": applications_by_reg,
    }
    with open(OUT_APPS, "w", encoding="utf-8") as f:
        json.dump(apps_output, f, ensure_ascii=False, separators=(",", ":"))
    apps_kb = OUT_APPS.stat().st_size / 1024

    print(f"\n=== 出力結果 ===")
    print(f"Main : {OUT_MAIN.name}  {main_kb:.1f} KB  ({main_kb/1024:.2f} MB)")
    print(f"Apps : {OUT_APPS.name}  {apps_kb:.1f} KB  ({apps_kb/1024:.2f} MB)")
    print(f"\nStats: {json.dumps(main_output['stats'], ensure_ascii=False, indent=2)}")

    # 失効剤のうちカテゴリ推定不能で除外したものを記録 (サイレント失敗防止)。
    # 殺そ剤・展着剤・植調剤など正規の対象外も含まれるため、件数のみで合否判定はしない。
    # 内容は監査用に毎回ファイルへ書き出し、ヒント表のカバレッジ漏れを追跡可能にする。
    dropped_path = RAW_DIR / "sikkou" / "unclassified_dropped.json"
    with open(dropped_path, "w", encoding="utf-8") as f:
        json.dump(
            [
                {"reg_no": c["reg_no"], "type_name": c["type_name"], "product_name": c["product_name"]}
                for c in cancelled_dropped
            ],
            f,
            ensure_ascii=False,
            indent=2,
        )
    print(f"\n失効剤カテゴリ推定不能で除外: {len(cancelled_dropped)} 件 (詳細: {dropped_path})")


if __name__ == "__main__":
    build()
